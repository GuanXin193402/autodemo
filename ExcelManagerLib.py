import openpyxl
import json
from urllib import parse
#from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment


class Excel_operation:
    def __init__(self, path, sheet_name):
        """
        这个是用来初始化读取对象的
        :param path: 文件路径 ---> str类型
        :param sheet_name: 表单名 ———> str类型
        """
        # 打开文件
        self.path = path
        self.wb = openpyxl.load_workbook(self.path)
        # 选择表单
        self.sheet = self.wb[sheet_name]

    # 将表格内容处理成想要的格式
    def readExcel(self):
        rows_data = list(self.sheet.rows)
        cases = []
        # 遍历所有行
        for case in rows_data[1:]:
            # 定义一个列表来存放各列值
            data = []
            # 获取各列单元格值并将其转为需要的格式加入列表中
            for cell in case[4:9]:
                # 判断单元格是否为空
                if cell.value:
                    # 将字典格式的字符串转换为字典：如请求头，请求体，期望返回码
                    if '{' in cell.value:
                        data.append(eval(cell.value))
                    else:
                        data.append(cell.value)
                else:
                    print('请确保请求相关参数不为空')
                    exit()
            if 'application/x-www-form-urlencoded' in data[2].values():
                data[3] = parse.urlencode(data[3]).replace('%27', '%22')
            else:
                data[3] = json.dumps(data[3])
            cases.append(data)
        # 关闭工作薄
        self.wb.close()
        return cases

    # 复制excel
    def copyExcel(self, report):
        self.wb.save(report)

    # 写入excel操作
    def writeExcel(self, case_row, result, reason=None):

        try:
            self.sheet.cell(row=case_row + 2, column=10, value=result)
            self.sheet.cell(row=case_row + 2, column=11, value=reason)
            #
            self.wb.save(self.path)
        except BaseException as e:
            raise e
        finally:
            self.wb.close()

    # 报告处理
    def optimizeReport(self):
        try:
            # red_fill = PatternFill("solid", fgColor="FF0000")
            rows = list(self.sheet.rows)
            count = 0
            for row in rows[1:]:
                # row[10].fill = red_fill if row[10].value == 'False' else count + 1
                if row[9].value == 'PASS':
                    count += 1
                else:
                    count += 0
            max_r = self.sheet.max_row
            self.sheet.cell(row=max_r + 1, column=1, value='成功率')
            self.sheet.cell(row=max_r + 1, column=2, value=round(count / (max_r - 1), 4) * 100)
            self.wb.save(self.path)
        except BaseException as e:
            raise e
        finally:
            self.wb.close()


if __name__ == '__main__':
    print(Excel_operation(r'E:\auto\auto_set\report\20200617测试结果.xlsx', '课程管理').optimizeReport())
